from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.title("Excel")
root.iconbitmap("favicon.ico")
root.geometry("300x300")


# Select Function
def select(e):
    my_label.config(text=my_listbox.get(ANCHOR))


# Create Listbox
my_listbox = Listbox(root, width=45)
my_listbox.pack(pady=20)

# Create a workbook
wb = load_workbook('Data.xlsx')

# Set active worksheet
ws = wb.active

# Grab a column of data
col_a = ws["A"]
col_b = ws["B"]

# Load list to listbox
for item in col_a:
    my_listbox.insert(END, item.value)

my_label = Label(root, text='Select Item...', font=("Helvetica", 18))
my_label.pack(pady=20)

# Create listbox binding
my_listbox.bind("<ButtonRelease-1>", select)

root.mainloop()